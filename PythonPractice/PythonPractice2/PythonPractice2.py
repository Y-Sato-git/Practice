#2020年6月分
flg =  input("YYYYMMDDの日付 または ALL ⇒ ")
print(flg)

if flg == "20200601" or flg == "ALL":

#2020/06/01
#例外処理
    l = [1,2,3,4,5]
    
    try:
        intNum = int(input("数字を入力してください　：　"))
        print(l[intNum + 1])
    except IndexError as ex:
        print("リストにありません。エラー内容⇒{}".format(ex))
    except Exception as ex:
        print("エラー内容⇒{}".format(ex))

if flg == "20200602" or flg == "ALL":

#2020/06/02
#defaultdict
    from collections import defaultdict

    strLine = input("文字列を入力してください　：　")

    #0で初期化
    d = defaultdict(int)

    for c in strLine:
        d[c] += 1
    print(d)

if flg == "20200603" or flg == "ALL":

#2020/06/03
#Class
    inName = input("Name : ")
    inAge = int(input("Age  : "))

    class Person(object):
        def name(self,strName):
            print(strName,"さん")
        def age(self,intAge):
            print(str(intAge),"歳")

    person = Person()
    person.name(inName)
    person.age(inAge)

if flg == "20200604" or flg == "ALL":

#2020/06/04
#ファイル(プロジェクトと同じフォルダに格納される)

    #書き込み
    f = open('test.txt','w')
    f.write('test')
    f.close()

    #読み込み
    f = open('test.txt','r')
    #3文字目(0から始まるので）
    f.seek(2)
    print(f.read(1))

'''
    以下のようにwithステートメントを使うと、自動でcloseしてくれる。
    with open('test.txt','w') as f:
        f.write('test')
'''

if flg == "20200605" or flg == "ALL":

#2020/06/05
#Excel自動化

    import openpyxl
    import docx

    strFileName = input("ファイル名：　")

    #workbookを開く
    wb = openpyxl.Workbook()

    #sheetを追加
    wb.create_sheet()

    #sheetの名前（一覧）
    print(wb.get_sheet_names())

    #sheet指定
    sheet = wb['Sheet']

    #cell指定
    c = sheet['A1']

    #cellの値変更
    c.value = "aaa"

    #2枚目
    sheet = wb['Sheet1']
    c = sheet['A1']
    c.value = "bbb"

    #保存
    wb.save(strFileName + ".xlsx")

if flg == "20200606" or flg == "ALL":

#2020/06/06
#Excel自動化続き,Word自動化

    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    sheet = wb.active
    _font = Font(name="明朝",sz=15,b=True)

    for i in range(1,100):
        cell = sheet.cell(row=i,column=1,value=i)
        cell.font = _font

    wb.save("20200606.xlsx")

    import docx
    from docx import Document

    document = Document()
    document.add_paragraph("AAAAAAAAAA")

    document.save("20200606.docx")

    import os
    print(os.listdir('.'))

if flg == "20200607" or flg == "20200608" or flg == "ALL":

#2020/06/07,08
#Tkinterお試し作成（別プロジェクト・消滅してた）
    print("今日は別プロジェクト")

if flg == "20200609" or flg == "ALL":

#2020/06/09
#logging
    import logging

    logging.basicConfig(level=logging.DEBUG)
    logging.debug('DEBUG')

if flg == "20200610" or flg == "ALL":

#2020/06/10
#openpyxl
    print("今日は別プロジェクト")

if flg == "20200611" or flg == "ALL":
    
#2020/06/11
#pandas
    import pandas
    import numpy

    array9 = numpy.array([1,2,3,4,5,6,7,8,9])
    dataframe9 = pandas.DataFrame({
        '1' : array9,
        '2' : array9 * 2,
        '3' : array9 * 3,
        '4' : array9 * 4,
        '5' : array9 * 5,
        '6' : array9 * 6,
        '7' : array9 * 7,
        '8' : array9 * 8,
        '9' : array9 * 9})
    print(dataframe9)

if flg == "20200612" or flg == "ALL":
    
#2020/06/12
#計算

    import scipy
    import numpy

    sample_data = numpy.array([1,2,3,4,5])
    print("sample_data = numpy.array([1,2,3,4,5])")
    print("sum  : " , str(scipy.sum(sample_data)))
    print("mean : " , str(scipy.mean(sample_data)))
    print("var(ddof=0) : " , str(scipy.var(sample_data,ddof=0)))
    print("var(ddof=1) : " , str(scipy.var(sample_data,ddof=1)))
    print("std(ddof=0) : " , str(scipy.std(sample_data,ddof=0)))
    print("amax : " , str(scipy.amax(sample_data)))
    print("amin : " , str(scipy.amin(sample_data)))
    print("median : " , str(scipy.median(sample_data)))

if flg == "20200613" or flg == "ALL":
    
#2020/06/13
#CSV
    import pandas

    csv_file = pandas.read_csv("TESTCSV.csv")
    print(csv_file)

if flg == "20200614" or flg == "ALL":
    
#2020/06/14
#グラフ
    import numpy
    import matplotlib.pyplot as mp

    x = numpy.linspace(0,10)
    line_1 = x
    line_2 = x * 2

    mp.xlabel("x",size = 20)
    mp.ylabel("y",size = 20)
    mp.title("グラフ")
    mp.grid()

    mp.plot(x,line_1)
    mp.plot(x,line_2,linestyle = "dashed")
    mp.legend()

    mp.show()

if flg == "20200615" or flg == "ALL":
    
#2020/06/15
#休養日
    print("手術日のためお休み")

if flg == "20200616" or flg == "ALL":
    
#2020/06/16
#paizaラーニング
    print("Classの勉強")

    name = input("名前を入力　：　")
    class Hello:
        def print_hello(self,name):
            print(name + "さん、こんにちは。")

    hello = Hello()
    hello.print_hello(name)
    
if flg == "20200617" or flg == "ALL":
    
#2020/06/17
#paizaラーニング
    print("2次元リストの勉強。リストの中にリストが入る。")

    list1 = ["a","b","c"]
    list2 = ["1","2","3"]
    list3 = ["A","B","C"]

    lists = [list1,list2,list3]
    print(lists)

if flg == "20200618" or flg == "ALL":
    
#2020/06/18
#paizaラーニング
    print("enumerate")

    animals = ["dog","cat","rabbit"]

    for (i,animal) in enumerate(animals):
        print(str(i + 1) + "番目は" + animal + "です。")

    #初期化
    numbers = [1 for i in range(10)]
    print()
    print("1で初期化")
    print(numbers)

if flg == "20200619" or flg == "ALL":
    
#2020/06/19
#paizaラーニング
    print("global変数")

    num = int(input("number : "))

    def print_num():
        global num
        print(num * 100)

    print_num()
    
if flg == "20200620" or flg == "ALL":
    
#2020/06/20
#paizaラーニング
    print("sorted")

    numList = [100,900,740,0,15]
    #昇順
    print(sorted(numList))
    #降順
    print(sorted(numList, reverse=True))

if flg == "20200621" or flg == "ALL":
    
#2020/06/21
#paizaラーニング
    print("関数のデフォルト値")

    def welcome(name = "noname"):
        print("こんにちは" + name + "さん")

    welcome()

if flg == "20200622" or flg == "ALL":
    
#2020/06/22
#paizaラーニング
    print("継承")

    greeting = input("Greeting : ")
    name = input("name     : ")

    class A:
        def __init__(self,greeting):
            self.greeting = greeting
            #print(greeting)

    class B(A):
        def say_hello(self,name):
            self.name = name
            print(self.greeting + " " + name)

    b = B(greeting)
    b.say_hello(name)
